import line_profiler
import time
import os
import psutil
import openpyxl
import numpy as np
from dolfinx import fem, mesh
from mpi4py import MPI
# MPI GATHER FUNCTION ---------------------------------
def mpi_gather(dfx_func, func_space, x0_con):
    # GET LOCAL RANGES AND GLOBAL SIZE OF ARRAY ---------------------------------
    imap = dfx_func.function_space.dofmap.index_map
    local_range = np.asarray(imap.local_range, dtype=np.int32) * dfx_func.function_space.dofmap.index_map_bs
    size_global = imap.size_global * dfx_func.function_space.dofmap.index_map_bs
    # GATHERED RANGES AND LOCAL DATA ---------------------------------
    gather_ranges = MPI.COMM_WORLD.gather(local_range, root=0)
    gather_data = MPI.COMM_WORLD.gather(dfx_func.vector.array, root=0)
    # COMMUNICATE LOCAL DOF COORDINATES ---------------------------------
    x = func_space.tabulate_dof_coordinates()[:imap.size_local]
    x_glob = MPI.COMM_WORLD.gather(x, root=0)
    # CREATE LOCAL DOF INDEXING ARRAY ---------------------------------
    local_dof_indexing = []
    for coord in x:
        local_dof_indexing.append(np.abs(x - coord).sum(axis=1).argmin())
    local_dof_size = MPI.COMM_WORLD.gather(len(local_dof_indexing), root=0)
    # DECLARE GATHERED PARALLEL ARRAYS ---------------------------------
    global_array = np.zeros(size_global)
    dfx_func_from_global = np.zeros(global_array.shape)
    serial_to_global, scatter_size = [], []
    # GATHER ---------------------------------
    if MPI.COMM_WORLD.rank == 0:
        # CREATE ARRAY RECEIVED ALL VALUES ---------------------------------
        for r, d in zip(gather_ranges, gather_data):
            global_array[r[0]:r[1]] = d
        # CREATE ARRAY WITH ALL COORDINATES ---------------------------------
        global_x = np.zeros((size_global, 3))
        for r, x_ in zip(gather_ranges, x_glob):
            global_x[r[0]:r[1], :] = x_
        for coord in x0_con:
            serial_to_global.append(np.abs(global_x - coord).sum(axis=1).argmin())
        # PREPARE SCATTER (DELETE END OF LOCAL DOF OF SIZE) ---------------------------------
        scatter_size = np.delete(local_dof_size, len(local_dof_size) - 1)
        # CREATE SORTED ARRAY FROM ---------------------------------
        for serial, glob in enumerate(serial_to_global):
            dfx_func_from_global[serial] = global_array[glob]
    return dfx_func_from_global, global_array, serial_to_global, scatter_size
# MPI SCATTER FUNCTION ---------------------------------
def mpi_gathered_scatter(dfx_func_from_global, global_array, serial_to_global, scatter_size):
    global_from_dfx_func, raveled = np.zeros(global_array.shape), []
    # SCATTER ---------------------------------
    if MPI.COMM_WORLD.rank == 0:
        # BACK TO GLOBAL ARRAY ---------------------------------
        for serial, glob in enumerate(serial_to_global):
            global_from_dfx_func[glob] = dfx_func_from_global[serial]
        # print(f"u_re to global array: \n{global_from_u}\n")
        split_global_array = np.array_split(global_from_dfx_func, np.cumsum(scatter_size))
        raveled = [np.ravel(arr_2) for arr_2 in split_global_array]
    scatter_global_array = MPI.COMM_WORLD.scatter(raveled, root=0)
    return scatter_global_array
# GENERAL MEMORY USAGE (PERCENT) ---------------------------------
def show_ram_percent():
    memory_usage_dict = dict(psutil.virtual_memory()._asdict())
    memory_usage_percent = memory_usage_dict['percent']
    return memory_usage_percent
# CURRENT PROCESS MEMORY USAGE (MB) ---------------------------------
def show_current_ram():
    pid = os.getpid()
    current_process = psutil.Process(pid)
    current_process_memory_usage = current_process.memory_info().rss / 2**20 # Bytes to MB
    return current_process_memory_usage
# IF WANT TO SEE LINE BY LINE PROFILING ---------------------------------
# (ACTIVE @profile, RUN THE COMMAND "mpirun -n 4 kernprof -v -l 2d_gather_scatter_func_profiler_write_2cor.py" and "python3 -m line_profiler 2d_gather_scatter_func_profiler_write_2cor.py.lprof > results.txt" in cmd)
@profile
def main():
    # INPUT PARAMETERS ---------------------------------
    nelx = 3
    nely = 3
    volfrac = 0.5
    # MESH ON CONDUCTOR PROCESS ---------------------------------
    if MPI.COMM_WORLD.rank == 0:
        msh_self = mesh.create_rectangle(MPI.COMM_SELF, ((0.0, 0.0), (nelx, nely)), (nelx, nely), cell_type=mesh.CellType.quadrilateral)
        U_self = fem.FunctionSpace(msh_self, ("DG", 0))
        x_self = U_self.tabulate_dof_coordinates()
        u_self = fem.Function(U_self)
        u_self.x.array[:] = volfrac
    else:
        x_self = None
    x_self = MPI.COMM_WORLD.bcast(x_self, root=0)
    # DECLARE MESH & FUNCTION ---------------------------------
    msh = mesh.create_unit_square(MPI.COMM_WORLD, nelx, nely, cell_type=mesh.CellType.quadrilateral)
    D0 = fem.FunctionSpace(msh, ("DG", 0))
    d = fem.Function(D0)
    # PARALLEL INTERPOLATION ---------------------------------
    d.x.array[:] = volfrac
    # print(f"Each local u array on process {MPI.COMM_WORLD.rank}: \n{d.vector.array}\n")
    # SET LOOP ---------------------------------
    loop, iter_t1, iter_t2 = 0, 0, 0
    if MPI.COMM_WORLD.rank == 2:
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1.title = "gather"
        sheet2 = wb.create_sheet("scatter")
    else:
        wb = None
        sheet1 = None
        sheet2 = None
    while loop < 2000:
        loop += 1
        # CHECK CPU, RAM USAGE BEFORE RUN GATHER FUNCTION ---------------------------------
        print(f"iter. {loop} step")
        before_usage1 = show_ram_percent()
        print(f"BEFORE GATHER CODE: memory usage {MPI.COMM_WORLD.rank}: {before_usage1}%")
        before_cur_pros_usage1 = show_current_ram()
        print(f"BEFORE GATHER CODE: current memory {MPI.COMM_WORLD.rank}: {before_cur_pros_usage1: 10.5f} MB")
        # MPI GATHER ---------------------------------
        start1 = time.time()
        u_from_global, global_array, serial_to_global, scatter_size = mpi_gather(d, D0, x_self)
        check_point1 = time.time()
        iter_t1 = iter_t1 + (check_point1 - start1)
        print(f"Gather time at iter.{loop}:", iter_t1, "sec")
        # CHECK CPU, RAM USAGE AFTER RUN GATHER FUNCTION ---------------------------------
        after_usage1 = show_ram_percent()
        print(f"AFTER GATHER CODE: memory usage {MPI.COMM_WORLD.rank}: {after_usage1}%")
        after_cur_pros_usage1 = show_current_ram()
        print(f"AFTER GATHER CODE: current memory {MPI.COMM_WORLD.rank}: {after_cur_pros_usage1: 10.5f} MB")
        print(f"RESIDUAL OF GATHER CODE {MPI.COMM_WORLD.rank}: {after_cur_pros_usage1 - before_cur_pros_usage1: 10.5f} MB")

        # MPI SCATTER ---------------------------------
        # CHECK CPU, RAM USAGE BEFORE RUN SCATTER FUNCTION ---------------------------------
        before_usage2 = show_ram_percent()
        print(f"BEFORE SCATTER CODE: memory usage {MPI.COMM_WORLD.rank}: {before_usage2}%")
        before_cur_pros_usage2 = show_current_ram()
        print(f"BEFORE SCATTER CODE: current memory {MPI.COMM_WORLD.rank}: {before_cur_pros_usage2: 10.5f} MB")
        start2 = time.time()
        scatter_global_array = mpi_gathered_scatter(u_from_global, global_array, serial_to_global, scatter_size)
        check_point2 = time.time()
        iter_t2 = iter_t2 + (check_point2 - start2)
        print(f"Scatter time at iter.{loop}:", iter_t2, "sec")
        # CHECK CPU, RAM USAGE AFTER RUN SCATTER FUNCTION ---------------------------------
        after_usage2 = show_ram_percent()
        print(f"AFTER SCATTER CODE: memory usage {MPI.COMM_WORLD.rank}: {after_usage2}%")
        after_cur_pros_usage2 = show_current_ram()
        print(f"AFTER SCATTER CODE: current memory {MPI.COMM_WORLD.rank}: {after_cur_pros_usage2: 10.5f} MB")
        print(f"RESIDUAL OF SCATTER CODE {MPI.COMM_WORLD.rank}: {after_cur_pros_usage2 - before_cur_pros_usage2: 10.5f} MB\n")
        # MPI SEND & RECEIVED DATA ---------------------------------
        if MPI.COMM_WORLD.rank == 0:
            MPI.COMM_WORLD.send(iter_t1, dest=2, tag=1)
            MPI.COMM_WORLD.send(before_cur_pros_usage1, dest=2, tag=3)
            MPI.COMM_WORLD.send(after_cur_pros_usage1, dest=2, tag=5)
            MPI.COMM_WORLD.send(iter_t2, dest=2, tag=6)
            MPI.COMM_WORLD.send(before_cur_pros_usage2, dest=2, tag=8)
            MPI.COMM_WORLD.send(after_cur_pros_usage2, dest=2, tag=10)
        elif MPI.COMM_WORLD.rank == 1:
            MPI.COMM_WORLD.send(iter_t1, dest=2, tag=11)
            MPI.COMM_WORLD.send(before_cur_pros_usage1, dest=2, tag=13)
            MPI.COMM_WORLD.send(after_cur_pros_usage1, dest=2, tag=15)
            MPI.COMM_WORLD.send(iter_t2, dest=2, tag=16)
            MPI.COMM_WORLD.send(before_cur_pros_usage2, dest=2, tag=18)
            MPI.COMM_WORLD.send(after_cur_pros_usage2, dest=2, tag=20)
        elif MPI.COMM_WORLD.rank == 3:
            MPI.COMM_WORLD.send(iter_t1, dest=2, tag=31)
            MPI.COMM_WORLD.send(before_cur_pros_usage1, dest=2, tag=33)
            MPI.COMM_WORLD.send(after_cur_pros_usage1, dest=2, tag=35)
            MPI.COMM_WORLD.send(iter_t2, dest=2, tag=36)
            MPI.COMM_WORLD.send(before_cur_pros_usage2, dest=2, tag=38)
            MPI.COMM_WORLD.send(after_cur_pros_usage2, dest=2, tag=40)
        elif MPI.COMM_WORLD.rank == 2:
            iter_t1_0 = MPI.COMM_WORLD.recv(source=0, tag=1)
            before_cur_pros_usage1_0 = MPI.COMM_WORLD.recv(source=0, tag=3)
            after_cur_pros_usage1_0 = MPI.COMM_WORLD.recv(source=0, tag=5)
            iter_t2_0 = MPI.COMM_WORLD.recv(source=0, tag=6)
            before_cur_pros_usage2_0 = MPI.COMM_WORLD.recv(source=0, tag=8)
            after_cur_pros_usage2_0 = MPI.COMM_WORLD.recv(source=0, tag=10)

            iter_t1_1 = MPI.COMM_WORLD.recv(source=1, tag=11)
            before_cur_pros_usage1_1 = MPI.COMM_WORLD.recv(source=1, tag=13)
            after_cur_pros_usage1_1 = MPI.COMM_WORLD.recv(source=1, tag=15)
            iter_t2_1 = MPI.COMM_WORLD.recv(source=1, tag=16)
            before_cur_pros_usage2_1 = MPI.COMM_WORLD.recv(source=1, tag=18)
            after_cur_pros_usage2_1 = MPI.COMM_WORLD.recv(source=1, tag=20)

            iter_t1_3 = MPI.COMM_WORLD.recv(source=3, tag=31)
            before_cur_pros_usage1_3 = MPI.COMM_WORLD.recv(source=3, tag=33)
            after_cur_pros_usage1_3 = MPI.COMM_WORLD.recv(source=3, tag=35)
            iter_t2_3 = MPI.COMM_WORLD.recv(source=3, tag=36)
            before_cur_pros_usage2_3 = MPI.COMM_WORLD.recv(source=3, tag=38)
            after_cur_pros_usage2_3 = MPI.COMM_WORLD.recv(source=3, tag=40)
        # WRITE GATHER DATA ---------------------------------
        if MPI.COMM_WORLD.rank == 2:
            sheet1.cell(row=1, column=1).value = "loop"
            sheet1.cell(row=1, column=2).value = "time"
            sheet1.cell(row=1, column=3).value = "before mem"
            sheet1.cell(row=1, column=4).value = "after mem"
            sheet1.cell(row=1, column=5).value = "residual"
            sheet1.cell(row=loop+1, column=1).value = loop
            sheet1.cell(row=loop+1, column=2).value = iter_t1_0
            sheet1.cell(row=loop+1, column=3).value = before_cur_pros_usage1_0
            sheet1.cell(row=loop+1, column=4).value = after_cur_pros_usage1_0
            sheet1.cell(row=loop+1, column=5).value = after_cur_pros_usage1_0 - before_cur_pros_usage1_0

            sheet1.cell(row=1, column=7).value = "loop"
            sheet1.cell(row=1, column=8).value = "time"
            sheet1.cell(row=1, column=9).value = "before mem"
            sheet1.cell(row=1, column=10).value = "after mem"
            sheet1.cell(row=1, column=11).value = "residual"
            sheet1.cell(row=loop + 1, column=7).value = loop
            sheet1.cell(row=loop + 1, column=8).value = iter_t1_1
            sheet1.cell(row=loop + 1, column=9).value = before_cur_pros_usage1_1
            sheet1.cell(row=loop + 1, column=10).value = after_cur_pros_usage1_1
            sheet1.cell(row=loop + 1, column=11).value = after_cur_pros_usage1_1 - before_cur_pros_usage1_1

            sheet1.cell(row=1, column=13).value = "loop"
            sheet1.cell(row=1, column=14).value = "time"
            sheet1.cell(row=1, column=15).value = "before mem"
            sheet1.cell(row=1, column=16).value = "after mem"
            sheet1.cell(row=1, column=17).value = "residual"
            sheet1.cell(row=loop + 1, column=13).value = loop
            sheet1.cell(row=loop + 1, column=14).value = iter_t1
            sheet1.cell(row=loop + 1, column=15).value = before_cur_pros_usage1
            sheet1.cell(row=loop + 1, column=16).value = after_cur_pros_usage1
            sheet1.cell(row=loop + 1, column=17).value = after_cur_pros_usage1 - before_cur_pros_usage1

            sheet1.cell(row=1, column=19).value = "loop"
            sheet1.cell(row=1, column=20).value = "time"
            sheet1.cell(row=1, column=21).value = "before mem"
            sheet1.cell(row=1, column=22).value = "after mem"
            sheet1.cell(row=1, column=23).value = "residual"
            sheet1.cell(row=loop + 1, column=19).value = loop
            sheet1.cell(row=loop + 1, column=20).value = iter_t1_3
            sheet1.cell(row=loop + 1, column=21).value = before_cur_pros_usage1_3
            sheet1.cell(row=loop + 1, column=22).value = after_cur_pros_usage1_3
            sheet1.cell(row=loop + 1, column=23).value = after_cur_pros_usage1_3 - before_cur_pros_usage1_3
        # WRITE SCATTER DATA ---------------------------------
        if MPI.COMM_WORLD.rank == 2:
            sheet2.cell(row=1, column=1).value = "loop"
            sheet2.cell(row=1, column=2).value = "time"
            sheet2.cell(row=1, column=3).value = "before mem"
            sheet2.cell(row=1, column=4).value = "after mem"
            sheet2.cell(row=1, column=5).value = "residual"
            sheet2.cell(row=loop + 1, column=1).value = loop
            sheet2.cell(row=loop + 1, column=2).value = iter_t2_0
            sheet2.cell(row=loop + 1, column=3).value = before_cur_pros_usage2_0
            sheet2.cell(row=loop + 1, column=4).value = after_cur_pros_usage2_0
            sheet2.cell(row=loop + 1, column=5).value = after_cur_pros_usage2_0 - before_cur_pros_usage2_0

            sheet2.cell(row=1, column=7).value = "loop"
            sheet2.cell(row=1, column=8).value = "time"
            sheet2.cell(row=1, column=9).value = "before mem"
            sheet2.cell(row=1, column=10).value = "after mem"
            sheet2.cell(row=1, column=11).value = "residual"
            sheet2.cell(row=loop + 1, column=7).value = loop
            sheet2.cell(row=loop + 1, column=8).value = iter_t2_1
            sheet2.cell(row=loop + 1, column=9).value = before_cur_pros_usage2_1
            sheet2.cell(row=loop + 1, column=10).value = after_cur_pros_usage2_1
            sheet2.cell(row=loop + 1, column=11).value = after_cur_pros_usage2_1 - before_cur_pros_usage2_1

            sheet2.cell(row=1, column=13).value = "loop"
            sheet2.cell(row=1, column=14).value = "time"
            sheet2.cell(row=1, column=15).value = "before mem"
            sheet2.cell(row=1, column=16).value = "after mem"
            sheet2.cell(row=1, column=17).value = "residual"
            sheet2.cell(row=loop + 1, column=13).value = loop
            sheet2.cell(row=loop + 1, column=14).value = iter_t2
            sheet2.cell(row=loop + 1, column=15).value = before_cur_pros_usage2
            sheet2.cell(row=loop + 1, column=16).value = after_cur_pros_usage2
            sheet2.cell(row=loop + 1, column=17).value = after_cur_pros_usage2 - before_cur_pros_usage2

            sheet2.cell(row=1, column=19).value = "loop"
            sheet2.cell(row=1, column=20).value = "time"
            sheet2.cell(row=1, column=21).value = "before mem"
            sheet2.cell(row=1, column=22).value = "after mem"
            sheet2.cell(row=1, column=23).value = "residual"
            sheet2.cell(row=loop + 1, column=19).value = loop
            sheet2.cell(row=loop + 1, column=20).value = iter_t2_3
            sheet2.cell(row=loop + 1, column=21).value = before_cur_pros_usage2_3
            sheet2.cell(row=loop + 1, column=22).value = after_cur_pros_usage2_3
            sheet2.cell(row=loop + 1, column=23).value = after_cur_pros_usage2_3 - before_cur_pros_usage2_3
        # FINAL RESULTS ---------------------------------
        d.vector.array = scatter_global_array
        if MPI.COMM_WORLD.rank == 2:
            wb.save('iter_time.xlsx')

if __name__ == "__main__":
    main()